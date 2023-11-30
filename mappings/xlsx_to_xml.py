from openpyxl import load_workbook
from yattag import Doc, indent
'''
		dataset
		/    	\
	   / 	  	 \
	Sheet_1  	Names
	/				\
letters_metadata	 Person
'''


# Load our Excel File
wb = load_workbook("fixed.xlsx")
# Getting an object of active sheet 1
ws = wb.worksheets[0]
#getting an object of sheet 3 with title 'Names' 
ws_names = wb.worksheets[2]
ws_subjects = wb.worksheets[1]
ws_images = wb.worksheets[3]

# match names with their respective id's
def get_ids(ws_names):
	names = []
	ids = [] # list for gathering ids of names listed in worksheet 1
	for column in ws_names.iter_cols():
		# Get the value of the first cell in the
		# column (the cell with the column name)
		column_name = column[0].value
		# Check if the column is the "Name" column
		if column_name == "Name (Last, First)":
			# Iterate over the cells in the column
			for cell in column:
				# Add the value of the cell to the list
				if cell.value != 0:
					names.append(cell.value)
		if column_name == "identifier":
			for cell in column:
			# Add the value of the cell to the list
				if cell.value != 0:
					ids.append(cell.value)
	ids = ids[:len(names)]
	id_dict={}
	id_dict['0'] = 0
	#create dictionary with keys the names and values the ids
	for i in range(len(names)):
		id_dict[names[i]] = ids[i]
	return id_dict

name_id_dict = get_ids(ws_names)

def get_subject_ids(ws_subjects):
	names = []
	ids = []
	i=0
	for column in ws_subjects.iter_cols():
		# Get the value of the first cell in the
		# column (the cell with the column name)
		column_name = column[0].value
		# Check if the column is the "Name" column
		if column_name == "name":
			# Iterate over the cells in the column
			for cell in column:
				i=i+1
				# Add the value of the cell to the list
				if cell.value != 0:
					names.append(cell.value)
					ids.append(i)
	ids = ids[:len(names)]
	id_dict={}
	id_dict['0'] = 0
	#create dictionary with keys the names and values the ids
	for i in range(len(names)):
		id_dict[names[i]] = ids[i]
	return id_dict

subjects_dict = get_subject_ids(ws_subjects)
  
# Print the list of names
#get tag for each column
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=29):
    clean_tags = [cell.value for cell in row]
#make tag suitable for xml format
tags = []
char_remov = [" ","(",")","/","?","-",",","'"]
for tag in clean_tags:
    for char in char_remov:
    # replace() "returns" an altered string
        tag = tag.replace(char, "_")
    tags.append(tag)
    
#tag names for Names sheet
for row in ws_names.iter_rows(min_row=1, max_row=1, min_col=1, max_col=5):
	clean_tag_names = [cell.value for cell in row]
tags_names = []
char_remov = [" ","(",")","/","?","-",",","'"]
for tag in clean_tag_names:
	for char in char_remov:
	# replace() "returns" an altered string
		tag = tag.replace(char, "_")
	tags_names.append(tag)


# Returning returns a triplet
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
# xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'
# Appends the String to document
doc.asis(xml_header)
# doc.asis(xml_schema)
id_num = 1000 #number for new id's
id_subject_num=100

with tag('dataset'):
#create xml with starting node Sheet_1
	with tag('Sheet1'):
		for row in ws.iter_rows(min_row=2, max_row=607, min_col=1, max_col=29):
			row = [cell.value for cell in row]
			#child node for Sheet_1 is letters_metadata
			with tag("letters_metadata"):
				for i in range(29):
					#childs of letters_metadata are each one of the names of the 29 columns
					#normalizations take place for columns: Subject_people, Subjects, letter_contents, Notes, Accompanying_Material and I_tatti_control num
					#which means we separate values that are separated by comma or semicolon
					#also we keep as childs for the person their names and ids in order to make properly their uri's
					# print(row[i],type(row[i]))
					if str(row[i]) != '0':
						if tags[i]=='Subject____People__Last_Name__First_Name_':
							names_splitted = str(row[i]).split(";")
							if names_splitted[-1] == '' or names_splitted[-1] == ' ':
								names_splitted = names_splitted[:-1]
							for j in range(len(names_splitted)):
								#due to split method, names in splitted list have a space as first character, except first name, so we need to remove it 
								if j == 0:
									name = names_splitted[j]
								else:
									name = names_splitted[j][1:]
								with tag(tags[i]):
									with tag('name'):
										text(name)
									with tag('id'):
										text(name_id_dict.get(name,id_num)) # get value of name if exists else id_num
										if name not in name_id_dict:
											name_id_dict[name] = id_num
											id_num += 1
											
						elif tags[i]=='Subjects':
							subjects_splitted = str(row[i]).split(";")
							if subjects_splitted[-1] == '' or subjects_splitted[-1] == ' ':
								subjects_splitted = subjects_splitted[:-1]
							for j in range(len(subjects_splitted)):
								#due to split method, names in splitted list have a space as first character, except first name, so we need to remove it
								with tag(tags[i]):
									if j ==0: 
										name = subjects_splitted[j]
									else:
										name = subjects_splitted[j][1:]
									# text(name)
									with tag('name'):
										text(name)
									with tag('id'):
										text(subjects_dict.get(name,id_subject_num)) # get value of name if exists else id_subject_num
										if name not in subjects_dict:
											subjects_dict[name] = id_subject_num
											id_subject_num += 1
						elif tags[i]=='I_Tatti_Control_Number_s_':
							numbers_splitted = str(row[i]).split(",")
							if numbers_splitted[-1] == '' or numbers_splitted[-1] == ' ':
								numbers_splitted = numbers_splitted[:-1]
							for j in range(len(numbers_splitted)):
								with tag(tags[i]):
									if j ==0: 
										text(numbers_splitted[j])
									else:
										text(numbers_splitted[j][0:])
						elif tags[i]=='Letter_Contents':
							contents_splitted = str(row[i]).split(";")
							if contents_splitted[-1] == '' or contents_splitted[-1] == ' ':
								contents_splitted = contents_splitted[:-1]
							for j in range(len(contents_splitted)):
								with tag(tags[i]):
									if j ==0: 
										text(contents_splitted[j])
									else:
										text(contents_splitted[j][0:])
						elif tags[i]=='Notes':
							contents_splitted = str(row[i]).split(";")
							if contents_splitted[-1] == '' or contents_splitted[-1] == ' ':
								contents_splitted = contents_splitted[:-1]
							for j in range(len(contents_splitted)):
								with tag(tags[i]):
									if j ==0: 
										text(contents_splitted[j])
									else:
										text(contents_splitted[j][0:])
						elif tags[i]=='Accompanying_Material':
							contents_splitted = str(row[i]).split(";")
							if contents_splitted[-1] == '' or contents_splitted[-1] == ' ':
								contents_splitted = contents_splitted[:-1]
							for j in range(len(contents_splitted)):
								with tag(tags[i]):
									if j ==0: 
										text(contents_splitted[j])
									else:
										text(contents_splitted[j][0:])
						# below we add sender's id in the block of sender in order to add it to the uri in next step
						elif tags[i]=='sender':
							name = row[i]
							with tag(tags[i]):
								with tag('name'):
									text(name)
								with tag('id'):
									#check if name is in dictionary with names and in this case take the known id, else store new name in dict
									text(name_id_dict.get(name,id_num))
									if name not in name_id_dict:
										name_id_dict[name] = id_num
										id_num += 1
						# below we add recipient's id in the block of recipient in order to add it to the uri in next step
						elif tags[i]=='recipient':
							name = row[i]
							with tag(tags[i]):
								with tag('name'):
									text(name)
								with tag('id'):
									#check if name is in dictionary with names and in this case take the known id, else store new name in dict
									text(name_id_dict.get(name,id_num))
									if name not in name_id_dict:
										name_id_dict[name] = id_num
										id_num += 1
						else:
							if tags[i] == "Name_of_Transcriber_1__Last_Name__First_Name_":
								with tag("Transcriber_ID_1"):
									text("1")
							if tags[i] == "Name_of_Transcriber_2__Last_Name__First_Name_":
								with tag("Transcriber_ID_2"):
									text("2")
							with tag(tags[i]):
								text(str(row[i]).strip())

	#create second node with start from Names
	with tag('Names'):
		for row in ws_names.iter_rows(min_row=2, max_row=676, min_col=1, max_col=5):
			row = [cell.value for cell in row]
			#child node for Sheet_1 is letters_metadata
			with tag("Person"):
				for i in range(5):
					with tag(tags_names[i]):
						text(row[i])
	#create third node with start from Subjects_art
	with tag('Subjects_art'):
		for row in ws_subjects.iter_rows(min_row=2, max_row=65, min_col=1, max_col=2):
			row = [cell.value for cell in row]
			with tag('subject_data'):
				with tag('name'):
					text(row[0])
				with tag('scope'):
					text(row[1])
				with tag('id'):
					text(subjects_dict[row[0]])
	with tag('Images'):
		for row in ws_images.iter_rows(min_row=2, max_row=606, min_col=1, max_col=33):
			values = [cell.value for cell in row]
			with tag('Image'):
				for idx, item in enumerate(values):
					if idx == 0:
						with tag('Letter_ID'):
							text(item)
					elif idx != 0 and item != 0:
						with tag('Image_ID'):
							text(item.strip())


result = indent(
	doc.getvalue(),
	indentation='   ',
	indent_text=False
)

with open("output_normalized.xml", "w") as f:
	f.write(result)
